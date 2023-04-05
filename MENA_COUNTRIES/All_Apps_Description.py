import pandas as pd
from google_play_scraper import app
import numpy as np
import datetime
from googletrans import Translator
import openpyxl
from openpyxl import Workbook

translator = Translator()
def playstore_desc_scrap(): 
    country_name = 'United Arab Emirates'
    country_code = 'ae'
    workbook = Workbook()
    workbook.save('C:/Users/kc/Desktop/Play_store_webscraping/MENA_COUNTRIES/APP_DETAILS/'+country_name+".xlsx")
    categories = ['ART_AND_DESIGN','AUTO_AND_VEHICLES','BEAUTY','BOOKS_AND_REFERENCE','BUSINESS','COMICS','COMMUNICATION','DATING','EDUCATION','ENTERTAINMENT','EVENTS','FINANCE','FOOD_AND_DRINK','HEALTH_AND_FITNESS','HOUSE_AND_HOME','LIBRARIES_AND_DEMO','LIFESTYLE','MAPS_AND_NAVIGATION','MEDICAL','MUSIC_AND_AUDIO','NEWS_AND_MAGAZINES','PARENTING','PERSONALIZATION','PHOTOGRAPHY','PRODUCTIVITY','SHOPPING','SOCIAL','SPORTS','TOOLS','WEATHER','TRAVEL_AND_LOCAL','VIDEO_PLAYERS']
    #df1 = pd.read_csv('C:/Users/kc/Desktop/Play_store_webscraping/7_COUNTRIES_SUB_LINKS/play_store_categories.csv')
    for ct in categories[:2]:
        row_data = [['title','descriptionHTML','summary','installs','minInstalls','realInstalls','score','ratings',
        'price','free','currency','inAppProductPrice','developer','developerId','developerEmail',
        'developerWebsite','developerAddress','privacyPolicy','genreId','icon','headerImage','screenshots','video',
        'videoImage','contentRating','adSupported','containsAds','released','updated','version','appId','url']]
        data = pd.read_excel('C:/Users/kc/Desktop/Play_store_webscraping/MENA_COUNTRIES/APP_LINKS/United Arab Emirates.xlsx',sheet_name=ct)
        print(data)
        print("category name >>>>> ",ct)
        for ap in data['lnk']:
            try:
                print("link >>>> ",ap)
                app_id = ap.split('id=')[-1]
                print("appid >>>>> ",app_id)
                # if app_id == 'NaN':
                #     i = i+1
                # print(i,' ',app_id)
                a_id = app_id+'&hl=en'
                result = app(a_id)
                row_data.append([translator.translate(result['title']).text,translator.translate(result['descriptionHTML']).text,translator.translate(result['summary']).text,result['installs'],
                result['minInstalls'],result['realInstalls'],result['score'],result['ratings'],
                result['price'],result['free'],result['currency'],result['inAppProductPrice'],
                result['developer'],result['developerId'],result['developerEmail'],result['developerWebsite'],result['developerAddress'],
                result['privacyPolicy'],result['genreId'],result['icon'],result['headerImage'],result['screenshots'],result['video'],result['videoImage'],
                result['contentRating'],result['adSupported'],result['containsAds'],result['released'],datetime.datetime.fromtimestamp(result['updated']).strftime("%d/%m/%Y"),
                result['version'],result['appId'],result['url']])
                print("app name >>>>> ",result['title'])
                print("app url >>>>> ",result['url'])
            except:
                print('except part-----------')
        workbook = openpyxl.open('C:/Users/kc/Desktop/Play_store_webscraping/MENA_COUNTRIES/APP_DETAILS/'+country_name+'.xlsx')
        sheet = workbook.create_sheet(ct)
        for row in row_data:
            sheet.append(row)
        workbook.save('C:/Users/kc/Desktop/Play_store_webscraping/MENA_COUNTRIES/APP_DETAILS/'+country_name+'.xlsx')
        
        # datafrm = pd.DataFrame(row_data[1:],columns=['title','descriptionHTML','summary','installs','minInstalls','realInstalls','score','ratings',
        # 'price','free','currency','inAppProductPrice','developer','developerId','developerEmail',
        # 'developerWebsite','developerAddress','privacyPolicy','genreId','icon','headerImage','screenshots','video',
        # 'videoImage','contentRating','adSupported','containsAds','released','updated','version','appId','url'])
        # print(datafrm)
        # datafrm.to_csv('C:/Users/kc/Desktop/Play_store_webscraping/JORDAN_COUNTRY/APP_DESCRIPTION/'+ct+'.csv')
        # workbook = openpyxl.open('C:/Users/kc/Desktop/Play_store_webscraping/JORDAN_COUNTRY/APP_DESCRIPTION/JORDAN.xlsx')
        # sheet = workbook.create_sheet(ct)
        # for row in row_data:
        #     sheet.append(row)
        # workbook.save('C:/Users/kc/Desktop/Play_store_webscraping/JORDAN_COUNTRY/APP_DESCRIPTION/JORDAN.xlsx')
        
        #print(desc1)
        # print('before sorting------------------------------------')
        #print(desc)
        # #df.sort_values(by='score',ascending=False,inplace=True)
        # print('after sorting-----------------------------')
        #datetime.datetime.fromtimestamp(desc['updated']).strftime("%d/%m/%Y")
        # for dt in range(len(desc['updated'])):
        #     desc['updated'][dt] = datetime.datetime.fromtimestamp(desc['updated'][dt]).strftime("%d/%m/%Y")
        #desc.drop(['descriptionHTML','summary','histogram','saleTime','originalPrice','saleText','offersIAP','genre','comments'],axis=1,inplace=True)
        
        # workbook.save('C:/Users/kc/Desktop/Play_store_webscraping/7_COUNTRIES_SUB_LINKS/APP_DETAILS/JORDAN.xlsx')
    
playstore_desc_scrap()
